# -*- coding: UTF-8 -*- 

import openpyxl
import sys
# from tqdm.notebook import tqdm


#########################################################
### IO list file (.xlsx) operation
#########################################################
def  parse_IO_list(file_path, instance_file_path, difinition_file_path, interface_file_path):
    
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    with open(instance_file_path, "w") as f1 :
        with open(difinition_file_path, "w") as f2:
            with open(interface_file_path, 'w') as f3:
        
                header = [cell.value for cell in sheet[1]]  
                ### check list completeness
                if len(header) != 5:
                    print("No enough or too much irrelevent information in the IO list!!! Please check the completeness of the IO file!")
                    sys.exit()            
                for field in header:
                    if field.strip() not in ["IO_Name", "Type", "Direction", "Width", "Connection"]:
                        print("incorrect information in the IO list!!! Please check the completeness of the IO file!")
                        sys.exit()  

                for i, row in enumerate(sheet.iter_rows(min_row = 2, values_only = True)):
                    ### the IO list is orchestrated in the following manner : 
                    #       IO Name; Type(Logic, bit, wire or reg); Direction(Input ot Output); Width(Integer); Interface_connect(corresponding connect in the interface of a module); 
                    #   Eg.   sys_clk_i             wire                       input                      1                   sys_clk           
                    for entry in row:
                        if entry == None:
                        ### if there is an incomplete row of the IO file, report and quit the program
                            print("missing information for parsing in a row of the IO list!!! Please check the completeness of the IO file!")
                            sys.exit()
                    print("row parsing is complete")
                    
                    ### assign entries to each corresponding variables
                    IO_Name, Type, Dir, Width, If_connect = row
                    
                    # for row in sheet.iter_rows(min_row=2, values_only=True):
                    #     # Extract values dynamically using the column indices
                    #     io_name = row[column_indices['IO Name']]
                    #     io_type = row[column_indices['Type']]
                    #     direction = row[column_indices['Direction']]
                    #     width = row[column_indices['Width']]
                    #     interface_connect = row[column_indices['Interface_connect']]
                    
                    ### generate formatted line for 1. module IO definition 2. instantiation
                    if i == sheet.max_row - 2:
                        IO_instance_format = f" .{IO_Name.strip()}( {If_connect.strip() if If_connect else ''}) \t"
                        if int(Width) == 1:
                            IO_definition_format = f"{Dir.strip()}  {Type.strip()} \t\t           \t {IO_Name.strip()} \t"
                            IO_interface_format = f" logic  \t\t           \t {IO_Name.strip()} \t;"
                        else:
                            IO_definition_format = f"{Dir.strip()}  {Type.strip()} \t\t[{int(Width)-1} : 0]\t {IO_Name.strip()} \t"
                            IO_interface_format = f" logic  \t\t[{int(Width)-1} : 0]\t {IO_Name.strip()} \t;"
                    else:
                        IO_instance_format = f" .{IO_Name.strip()}( {If_connect.strip() if If_connect else ''})     ,\t"    
                        if int(Width) == 1:
                            IO_definition_format = f"{Dir.strip()}  {Type.strip()} \t\t           \t {IO_Name.strip()}     ,\t"
                            IO_interface_format = f" logic  \t\t           \t {IO_Name.strip()} \t;"
                        else:
                            IO_definition_format = f"{Dir.strip()}  {Type.strip()} \t\t[{int(Width)-1} : 0]\t {IO_Name.strip()}     ,\t"  
                            IO_interface_format = f" logic  \t\t[{int(Width)-1} : 0]\t {IO_Name.strip()} \t;"                             
                    
                    ### write txt files
                    f1.write(IO_instance_format+"\n")
                    f2.write(IO_definition_format+"\n")
                    f3.write(IO_interface_format+"\n")
        
    print("TXT files have been created successfully!")


if __name__ == "__main__":
    
    ### specify file paths
    file_path = "./IO_List.xlsx"
    instance_file_path = "./instance_temp.txt"
    difinition_file_path = "./IOdefinition_temp.txt"
    interface_file_path = "./interface_temp.txt"
    
    parse_IO_list(file_path, instance_file_path, difinition_file_path, interface_file_path)
    
    
    
    
    
            
                
                                
        
            
            
    
    

